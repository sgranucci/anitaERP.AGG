#!/usr/bin/env python3
"""Drill ERP vs Excel EFE Datos por concepto/asiento/cuenta."""
import json
import subprocess
import sys
from collections import defaultdict

EXCEL = '/home/sergio/tmp/Efe Anita BSA 31.05.26.xlsx'
REF = '/var/www/html/anitaERP/storage/templates/contable/efe_referencia_biy_mayo2026.json'
ERP_DIR = '/var/www/html/anitaERP'
CONCEPTOS = [7, 12, 13, 24, 55, 65, 45, 51, 47, 20, 5]

import openpyxl

def num(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return 0.0

def parse_concepto(s):
    parts = str(s).split()
    for i, p in enumerate(parts):
        if p == 'Concepto:' and i + 1 < len(parts) and parts[i + 1].isdigit():
            return int(parts[i + 1])
    return None

def load_excel(conceptos):
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb['Datos']
    by_c = {c: defaultdict(lambda: {'n': 0, 'neto': 0.0}) for c in conceptos}
    lines = {c: [] for c in conceptos}
    for r in range(7, 35000):
        cid = parse_concepto(ws.cell(r, 2).value)
        if cid not in conceptos:
            continue
        cu = str(ws.cell(r, 3).value or '').strip()
        if cu.startswith('Concepto:'):
            continue
        o, p = num(ws.cell(r, 15).value), num(ws.cell(r, 16).value)
        net = p - o
        asiento = int(ws.cell(r, 6).value or 0)
        desc = str(ws.cell(r, 11).value or '')[:50]
        tc = str(ws.cell(r, 8).value or '')
        key = f"{asiento}|{cu}"
        by_c[cid][key]['n'] += 1
        by_c[cid][key]['neto'] += net
        lines[cid].append({
            'asiento': asiento, 'cuenta': cu, 'neto': net, 'desc': desc, 'tc': tc,
            'pagos': o, 'cobros': p,
        })
    return by_c, lines

def load_erp(conceptos):
    cids = ','.join(map(str, conceptos))
    php = f'''
require 'vendor/autoload.php';
$app = require 'bootstrap/app.php';
$app->make(Illuminate\\Contracts\\Console\\Kernel::class)->bootstrap();
App\\Support\\Contable\\MayorConcepto\\MayorConceptoRuntimeSupport::elevarLimites();
ini_set('memory_limit','-1');
$svc = app(App\\Services\\Contable\\EfeMensualReporteService::class);
$r = $svc->generarDesdeFiltros(['empresa_id'=>1,'mes'=>5,'anio'=>2026,'moneda_id'=>1,'solo_moneda_origen'=>false]);
$permitidos = [{cids}];
$out = [];
foreach ($r['filas_datos'] ?? [] as $f) {{
    $cid = (int)($f['concepto_id'] ?? 0);
    if (!in_array($cid, $permitidos, true)) continue;
    $cu = (string)($f['cuenta_codigo'] ?? $f['cuenta'] ?? '');
    if (!str_contains($cu, '-')) {{
        $cu = substr($cu,0,6).'-'.substr($cu,6);
    }}
    $pagos = (float)($f['pagos'] ?? 0);
    $cobros = (float)($f['cobros'] ?? 0);
    $out[] = [
        'concepto_id' => $cid,
        'asiento' => (int)($f['nro_asiento'] ?? 0),
        'cuenta' => $cu,
        'desc' => mb_substr((string)($f['descripcion'] ?? ''), 0, 50),
        'tipo' => (string)($f['tipo_comp'] ?? ''),
        'pagos' => $pagos,
        'cobros' => $cobros,
        'neto' => $cobros - $pagos,
    ];
}}
echo json_encode($out, JSON_UNESCAPED_UNICODE);
'''
    proc = subprocess.run(['php', '-r', php], cwd=ERP_DIR, capture_output=True, text=True, timeout=600)
    if proc.returncode != 0:
        print(proc.stderr, file=sys.stderr)
        sys.exit(1)
    rows = json.loads(proc.stdout)
    by_c = {c: defaultdict(lambda: {'n': 0, 'neto': 0.0}) for c in conceptos}
    lines = {c: [] for c in conceptos}
    for row in rows:
        cid = row['concepto_id']
        key = f"{row['asiento']}|{row['cuenta']}"
        by_c[cid][key]['n'] += 1
        by_c[cid][key]['neto'] += row['neto']
        lines[cid].append(row)
    return by_c, lines

def main():
    ref = json.load(open(REF))
    excel_by, excel_lines = load_excel(CONCEPTOS)
    erp_by, erp_lines = load_erp(CONCEPTOS)

    def ref_val(c):
        return float(ref['resumen_pagos_col_b'].get(str(c), ref['resumen_pagos_col_b'].get(c, 0)))

    for cid in sorted(CONCEPTOS, key=lambda c: -abs(
        sum(v['neto'] for v in erp_by[c].values()) - ref_val(c)
    )):
        excel_total = sum(v['neto'] for v in excel_by[cid].values())
        erp_total = sum(v['neto'] for v in erp_by[cid].values())
        ref_total = float(ref['resumen_pagos_col_b'].get(str(cid), ref['resumen_pagos_col_b'].get(cid, 0)))
        diff = erp_total - ref_total
        if abs(diff) < 500:
            continue

        print(f"\n{'='*72}\nCONCEPTO {cid}  Excel ref {ref_total:,.2f}  ERP {erp_total:,.2f}  Δ {diff:,.2f}")

        keys = set(excel_by[cid]) | set(erp_by[cid])
        ranked = sorted(keys, key=lambda k: abs(excel_by[cid][k]['neto'] - erp_by[cid][k]['neto']), reverse=True)
        print(f"{'Asiento|Cuenta':<28} {'Excel':>14} {'ERP':>14} {'Δ':>12}")
        for k in ranked[:8]:
            en = excel_by[cid][k]['neto']
            rn = erp_by[cid][k]['neto']
            d = rn - en
            if abs(d) < 0.01:
                continue
            print(f"{k:<28} {en:>14,.2f} {rn:>14,.2f} {d:>12,.2f}")

        # asientos solo ERP en c55 214010-004
        if cid == 55:
            excel_as = {int(k.split('|')[0]) for k in excel_by[cid] if k.endswith('214010-004')}
            erp_as = {int(k.split('|')[0]) for k in erp_by[cid] if k.endswith('214010-004')}
            solo_erp = sorted(erp_as - excel_as)[:5]
            if solo_erp:
                print("  Muestra asientos ERP-only en 214010-004:")
                for row in erp_lines[cid]:
                    if row['asiento'] in solo_erp and row['cuenta'] == '214010-004':
                        print(f"    as={row['asiento']} net={row['neto']:,.2f} {row['desc'][:45]}")
                        break

        # líneas Excel-only c13 117010
        if cid == 13:
            excel_only = [l for l in excel_lines[cid] if l['cuenta'] == '117010-001']
            erp_as_117 = {l['asiento'] for l in erp_lines[cid] if l['cuenta'] == '117010-001'}
            missing = [l for l in excel_only if l['asiento'] not in erp_as_117]
            if missing:
                print("  Cheques 117010-001 en Excel, ausentes ERP:")
                for l in missing[:6]:
                    print(f"    as={l['asiento']} net={l['neto']:,.2f} {l['desc']}")

if __name__ == '__main__':
    main()
